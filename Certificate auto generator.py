import os
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from datetime import datetime


# Define paths
EXCEL_FILE_PATH = "C:\\Users\\Dell\\Desktop\\Hexamad\\hexamad.xlsx" 
TEMPLATE_PATH = "C:\\Users\\Dell\\Desktop\\Hexamad\\Sample certificate.png"  #Path to the certificate template image
OUTPUT_FOLDER = "C:\\Users\\Dell\\Desktop\\Certificates"  #Path where the certificates will be saved

# Ensure the output folder exists
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Load the Excel workbook
try:
    workbook = load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active
except FileNotFoundError:
    print(f"Error: The Excel file at {EXCEL_FILE_PATH} was not found.")
    exit(1)

# Customize font settings
FONT_PATH = "C:\\Users\\Dell\\Downloads\\Roboto\\Roboto-Black.ttf"
FONT_SIZE = 38  # Adjust font size as needed

# Check if font file exists
if not os.path.exists(FONT_PATH):
    print(f"Error: Font file at {FONT_PATH} not found.")
    exit(1)

font = ImageFont.truetype(FONT_PATH, FONT_SIZE)

# Iterate through the Excel sheet rows
for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is headers
    title, name, date = row

# Convert the date from month name to month number
    try:
        if isinstance(date, str):
            # Parse date with month name (e.g., "December 21, 2023")
            parsed_date = datetime.strptime(date, "%B %d, %Y")
        elif isinstance(date, datetime):
            parsed_date = date
        else:
            raise ValueError("Date format is unsupported")

        # Format the date as "dd-mm-yyyy" with month number
        formatted_date = parsed_date.strftime("%d-%m-%Y")
    except ValueError:
        formatted_date = str(date)  # Fallback to raw string if parsing fails

    # Open the template image
    try:
        with Image.open(TEMPLATE_PATH) as template:
            draw = ImageDraw.Draw(template)

            # Add text to the template (adjust positions as needed)
            draw.text((320, 460), f"{title} {name}", fill="black", font=font)
            draw.text((650, 640), f"{formatted_date}", fill="black", font=font)

            # Save the generated certificate
            output_path = os.path.join(OUTPUT_FOLDER, f"{name}_certificate.png")
            template.save(output_path)

            print(f"Certificate generated for {name} and saved to {output_path}")

    except FileNotFoundError:
        print(f"Error: Template image at {TEMPLATE_PATH} not found.")
        exit(1)
print("Certificate generation complete.")

